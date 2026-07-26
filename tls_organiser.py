#!/usr/bin/env python3

"""
tls-organiser

Author: Charalampos Spanias (mollysec)

Organises SSL/TLS-related scanner findings into reporting-friendly root cause categories.
"""

import argparse
import warnings
from collections import defaultdict
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

#---------------------------------------
# Constants
#---------------------------------------

TLS_FINDINGS = {
    "Deprecated SSL Support": {"ids": {"20007","78447","78479","89058"}},
    "Deprecated TLS Support": {"ids": {"104743","157288"}},
    "Weak Cipher Suites": {"ids": {"26928","65821","42873","81606"}},
    "Invalid TLS Certificate Configuration": {"ids": {"51192","57582","45410","45411","15901","56284"}},
    "Weak Certificate Cryptography": {"ids": {"35291","69551","60108","86067"}},
    "Weak DH Parameters (Logjam)": {"ids": {"83875"}},
    "Anonymous Cipher Suites Supported": {"ids": {"31705"}}
    }


CERT_ISSUES = {

    "57582": {
        "title": "Self-Signed Certificate",
        "commentary": 
            "The service presented a self-signed TLS certificate that was not issued by a trusted Certificate Authority (CA). As a result, "
            "clients could not establish a trusted chain of trust when connecting to the service and may receive warnings indicating that the certificate issuer is not trusted.\n\n"

            "In external environments, this weakens the trust model of TLS by removing independent verification of the server's identity. In internal environments, users"
            " and systems may become accustomed to bypassing certificate warnings, reducing the effectiveness of certificate validation as a security control. An attacker in an "
            "adversary-in-the-middle (AiTM) position could potentially abuse this condition by presenting a similarly untrusted certificate, increasing the risk of service impersonation.",
    },

    "15901": {
        "title": "Certificate Expiry",
        "commentary":
            "The TLS certificate presented by the service had expired and was no longer within its validity period. While encrypted communications may still be established, clients "
            "can no longer verify the authenticity of the service using the certificate, reducing the trust assurances normally provided by TLS.\n\n"

            "Expired certificates commonly result in browser or application warnings that users may ignore in order to access the service. Over time, this can "
            "encourage poor security practices and reduce confidence in certificate validation controls. Additionally, the use of expired certificates may indicate weaknesses "
            "in certificate lifecycle management and renewal processes.",
    },

    "45410": {
        "title": "Certificate Hostname Mismatch",
        "commentary":
            "The TLS certificate presented by the service contained a Common Name (CN) and/or Subject Alternative Name (SAN) that did not match the hostname used to access the service. "
            "Consequently, clients were unable to verify that the certificate belonged to the intended endpoint.\n\n"

            "Although communications may remain encrypted, the inability to properly validate server identity weakens the trust model of TLS. Users encountering certificate warnings "
            "may be unable to distinguish between a benign configuration issue and a potential adversary-in-the-middle (AitM) attack, increasing the likelihood that warnings are "
            "ignored and malicious endpoints are trusted.",
    },

    "45411": {
        "title": "Certificate Hostname Mismatch",
        "commentary":
            "The TLS certificate presented by the service contained a Common Name (CN) and/or Subject Alternative Name (SAN) that did not match the hostname used to access the service. "
            "Consequently, clients were unable to verify that the certificate belonged to the intended endpoint.\n\n"

            "Although communications may remain encrypted, the inability to properly validate server identity weakens the trust model of TLS. Users encountering certificate warnings "
            "may be unable to distinguish between a benign configuration issue and a potential adversary-in-the-middle (AitM) attack, increasing the likelihood that warnings are "
            "ignored and malicious endpoints are trusted.",
    },

    "51192": {
        "title": "Invalid or Untrusted Certificate Chain",
        "commentary":
            "The TLS certificate could not be fully validated because the certificate chain was incomplete, incorrectly configured, or issued by an untrusted authority. As a result, "
            "clients were unable to establish trust in the authenticity of the service despite successful encryption of communications.\n\n"

            "Where certificate trust cannot be established, users and applications may be forced to bypass validation warnings to access the service. This weakens the protection "
            "provided by TLS and increases the risk of service impersonation or adversary-in-the-middle attacks.",
    },

    "56284": {
        "title": "Certificate Basic Constraints / Key Usage Misconfiguration",
        "commentary":
            "The TLS certificate presented by the service did not adhere to recommended certificate profile requirements relating to certificate constraints or key usage extensions. "
            "These extensions define how a certificate may be used and help clients determine whether the certificate is appropriate for a particular purpose, such as server "
            "authentication or certificate signing.\n\n"

            "Incorrect or missing constraints can lead to certificate validation inconsistencies and may reduce confidence in the trustworthiness of the certificate. While modern "
            "clients often enforce their own validation behaviour, non-standard certificate profiles can introduce compatibility issues and weaken the overall assurance provided "
            "by the public key infrastructure (PKI).\n\n"

            "Although exploitation is unlikely in isolation, the issue indicates weak certificate management practices and may contribute to broader trust validation problems when "
            "combined with other certificate-related weaknesses.",
    }
}


REFERENCES = {

    "tls_best_practices":
        "https://github.com/ssllabs/research/wiki/ssl-and-tls-deployment-best-practices",

    "microsoft_tls":
        "https://learn.microsoft.com/en-us/dotnet/core/extensions/sslstream-best-practices",

    "ncsc_tls":
        "https://www.gov.uk/government/publications/email-security-standards/transport-layer-security-tls",

    "cipher_hardening":
        "https://oneuptime.com/blog/post/2026-03-20-disable-weak-tls-cipher-suites/view#step-4-ciphers-to-explicitly-disable",

    "nginx_ssl":
        "https://nginx.org/en/docs/http/ngx_http_ssl_module.html",

    "apache_ssl":
        "https://httpd.apache.org/docs/2.4/mod/mod_ssl.html",

    "iis_ssl":
        "https://techcommunity.microsoft.com/blog/iis-support-blog/how-to-disable-ssl-2-0-or-ssl-3-0-from-iis-server/287812",

    "iis_crypto":
        "https://www.nartac.com/Products/IISCrypto",

    "rc4":
        "https://support.microsoft.com/en-gb/topic/microsoft-security-advisory-update-for-disabling-rc4-479fd6f0-c7b5-0671-975b-c45c3f2c0540#ID0EFR",

    "cbc":
        "https://learn.microsoft.com/en-us/dotnet/standard/security/vulnerabilities-cbc-mode",

    "cipher_suite_info":
        "https://ciphersuite.info/search/?q=CBC",

    "freak":
        "https://freakattack.com/",

    "logjam_ms":
        "https://docs.microsoft.com/en-us/security-updates/securitybulletins/2015/ms15-055",

    "logjam":
        "https://weakdh.org/sysadmin.html"
}


TLS_LIBRARY = {

    "Deprecated SSL Support": {
        "title": "Deprecated SSL Version 2 and SSL Version 3 Support",

        "commentary":
            "The service supported legacy SSL protocols (SSL 2.0 and/or SSL 3.0), which contain well-established cryptographic "
            "weaknesses and are no longer considered secure. These protocols lack many of the security protections present in modern "
            "versions of TLS and have been superseded by more robust encryption standards.\n\n"
            
            "The continued availability of SSL 2.0 and SSL 3.0 exposes the service to a number of well-known attacks. SSL 3.0 is affected "
            "by vulnerabilities such as POODLE, which exploit weaknesses in the handling of block cipher padding and may allow portions of "
            "encrypted communications to be decrypted under specific conditions. Similarly, SSL 2.0 support may introduce exposure to attacks "
            "such as DROWN, which leverage weaknesses in SSL 2.0 to undermine the security of otherwise stronger TLS communications.\n\n"

            "Although modern clients typically negotiate newer TLS versions where available, support for legacy protocols increases the risk "
            "of protocol downgrade attacks. An attacker in an adversary-in-the-middle (AiTM) position may be able to influence protocol "
            "negotiation and force the use of weaker encryption mechanisms, reducing the overall effectiveness of transport layer security controls.\n\n"

            "Successful exploitation generally requires the ability to intercept and manipulate network traffic, and in some cases the presence of "
            "additional environmental conditions. Nevertheless, retaining support for SSL 2.0 and SSL 3.0 unnecessarily expands the attack surface, "
            "weakens the overall cryptographic posture of the service, and maintains exposure to protocol versions that are no longer required in "
            "modern environments."
        ,

        "solution":
            "Disable support for SSL 2.0 and SSL 3.0 on all affected services and restrict communications to modern TLS protocol versions, "
            "preferably TLS 1.2 or TLS 1.3. In addition:\n"
                "- Review all externally and internally exposed services for legacy protocol support.\n"
                "- Ensure that only secure TLS versions are permitted during protocol negotiation.\n"
                "- Replace any legacy clients or applications that require SSL 2.0 or SSL 3.0.\n"
                "- Review certificate and key reuse across services to eliminate potential cross-protocol attack scenarios.\n"
                "- Verify that strong cipher suites and modern key exchange mechanisms are used following protocol upgrades."
        ,

        "references": [
            "tls_best_practices",
            "microsoft_tls",
            "ncsc_tls",
            "cipher_hardening",
            "nginx_ssl",
            "apache_ssl",
            "iis_ssl",
            "iis_crypto"
        ]
    },

    "Deprecated TLS Support": {
        "title": "Deprecated TLS 1.0 and TLS 1.1 Protocol Support",
                
        "commentary":
            "The service supported TLS 1.0 and TLS 1.1, both of which are deprecated protocol versions that are no longer aligned with "
            "modern cryptographic standards. These protocols contain known weaknesses and lack a number of security enhancements introduced "
            "in later versions, such as improved cipher suite support and stronger protections against cryptographic attacks.\n\n"

            "Although modern clients will typically negotiate TLS 1.2 or higher where available, the continued support for TLS 1.0 and "
            "TLS 1.1 introduces the risk of protocol downgrade attacks. An attacker in a privileged network position may be able to force "
            "connections to use weaker encryption standards, reducing the overall effectiveness of transport layer security protections.\n\n"

            "Continued support for these legacy protocols unnecessarily expands the attack surface and weakens the overall security posture "
            "of the service."
            ,
                
        "solution":
            "Remove support for TLS 1.0 and TLS 1.1 and restrict communication to modern TLS versions, preferably TLS 1.2 or higher with "
            "secure cipher suites."
        ,
        
        "references": [
            "tls_best_practices",
            "microsoft_tls",
            "ncsc_tls",
            "cipher_hardening",
            "nginx_ssl",
            "apache_ssl",
            "iis_ssl",
            "iis_crypto"
        ]
    },

    "Weak Cipher Suites": {

        "title": "Weak and Deprecated Ciphers Suites Supported",
        
        "commentary":
            "The service supported weak and deprecated SSL/TLS cipher suites that no longer meet modern cryptographic standards. These "
            "cipher suites rely on outdated algorithms or cryptographic constructions that provide reduced protection compared to "
            "contemporary alternatives, weakening the overall confidentiality and integrity of encrypted communications.\n\n"

            "The identified cipher suites included legacy algorithms such as RC4, 3DES, and export-grade cryptographic suites designed "
            "around intentionally weakened cryptography. RC4-based cipher suites are affected by well-documented cryptographic weaknesses "
            "that may allow portions of encrypted communications to be recovered under specific conditions. Similarly, 3DES-based cipher "
            "suites are vulnerable to attacks such as SWEET32, which exploit the limited block size used by the cipher and may permit "
            "recovery of information from long-lived encrypted sessions.\n\n"

            "Although modern clients generally prioritise stronger cipher suites during TLS negotiation, the continued availability of "
            "weaker alternatives introduces the potential for downgrade or negotiation-based attacks. An attacker in an adversary-in-the-middle "
            "(AiTM) position may be able to influence the handshake process and force the use of weaker cryptographic options under certain conditions.\n\n"

            "Successful exploitation typically requires additional preconditions, including the ability to intercept network traffic and "
            "influence protocol negotiation. Nevertheless, the presence of weak cipher suites unnecessarily expands the attack surface and "
            "reduces the overall strength of the service's cryptographic configuration."
            ,
        
        "solution":
            "Disable support for weak and deprecated cipher suites and restrict the service to modern cryptographic algorithms. More specifically:\n"
                "- Remove support for RC4-based, 3DES-based, EXPORT_RSA, and other deprecated or weak cipher suites.\n"
                "- Prefer modern cipher suites supporting authenticated encryption, such as AES-GCM or ChaCha20-Poly1305.\n"
                "- Prefer cipher suites that provide forward secrecy through ECDHE key exchange.\n"
                "- Periodically review TLS configurations to ensure deprecated algorithms are removed as standards evolve."
        ,

        "references": [
            "rc4",
            "cbc",
            "cipher_suite_info",
            "cipher_hardening",
            "nginx_ssl",
            "apache_ssl",
            "iis_ssl",
            "iis_crypto",
            "freak"
        ]
    },

    "Weak DH Parameters (Logjam)": {

        "title": "Weak Diffie-Hellman Parameters Supported (Logjam)",

        "commentary":
            "The service supported weak Diffie-Hellman (DH) key exchange parameters with a modulus size of 1024 bits or less. "
            "Such parameters no longer meet modern cryptographic standards and are associated with attacks such as Logjam, "
            "which exploit weaknesses in commonly used or precomputed DH groups.\n\n"

            "Although modern clients typically negotiate the strongest available cryptographic settings, the use of weak DH parameters "
            "reduces the overall strength of the key exchange process and may weaken the confidentiality of encrypted communications. "
            "An attacker in an adversary-in-the-middle (AiTM) position may be able to influence negotiation and leverage these weaker "
            "parameters under specific conditions.\n\n"

            "Successful exploitation generally requires network positioning, favourable cryptographic conditions, and significant "
            "computational resources. Nevertheless, the continued use of weak DH parameters unnecessarily expands the attack surface "
            "and weakens the overall cryptographic posture of the affected service."
        ,

        "solution":
            "Replace weak Diffie-Hellman (DH) parameters with modern key exchange mechanisms and cryptographic standards. More specifically:\n"
                "- Configure affected services to use DH parameters with a modulus size of at least 2048 bits.\n"
                "- Where possible, migrate to Elliptic Curve DH Ephemeral (ECDHE) key exchange mechanisms.\n"
                "- Remove legacy or commonly shared DH groups.\n"
                "- Generate custom DH parameters where traditional DH key exchange remains required.\n"
                "- Periodically review TLS configurations to ensure deprecated key exchange mechanisms are removed and only "
                "approved cryptographic standards are supported."
        ,
                
        "references": [
            "logjam_ms",
            "logjam"
        ]
    },

        "Weak Certificate Cryptography": {
            
            "title": "Weak Certificate Cryptography",
            
            "commentary":
                "The TLS certificate chain presented by the service utilised cryptographic mechanisms that no longer "
                "align with current industry standards. Specifically, the certificate was signed using a deprecated "
                "hashing algorithm and/or contained RSA keys below the recommended minimum key length of 2048 bits.\n\n"

                "Weak certificate signature algorithms, such as SHA-1, are no longer considered sufficiently resistant "
                "to modern cryptographic attacks. Although practical exploitation is difficult, known weaknesses reduce "
                "confidence in the integrity of the certificate and weaken the trust assurances provided by TLS.\n\n"

                "Similarly, RSA keys shorter than 2048 bits provide a reduced security margin against advances in computational "
                "capabilities. While such keys are not considered practically breakable in most environments, their continued use "
                "reflects an outdated cryptographic configuration and does not meet current security best practices.\n\n"

                "Modern browsers, operating systems, and security standards increasingly reject or discourage the use of weak "
                "certificate algorithms and key lengths. As a result, continued reliance on deprecated cryptographic mechanisms "
                "may lead to interoperability issues, reduced trust in affected services, and a weakened overall cryptographic posture."
                ,
            
                "solution":
                    "Replace affected certificates with certificates that utilise modern cryptographic standards. Specifically:\n"
                        "- Replace certificates signed using deprecated hashing algorithms (e.g. SHA-1 or MD5) with certificates "
                        "signed using SHA-256 or stronger algorithms.\n"
                        "- Replace any certificate chain components that use RSA keys shorter than 2048 bits.\n"
                        "- Ensure certificate issuance and renewal processes enforce minimum cryptographic standards and "
                        "approved algorithms.\n"
                        "- Periodically review certificate inventories to identify deprecated algorithms or key lengths before "
                        "they become unsupported by clients and browsers."
                ,

                "references": [

                ]
        },

        "Anonymous Cipher Suites Supported": {
            
            "title": "Anonymous Cipher Suites Supported",
            
            "commentary":
                "The service supported anonymous SSL/TLS cipher suites (aNULL). These cipher suites provide encryption but do not "
                "perform server authentication, preventing clients from verifying the identity of the remote service.\n\n"

                "As a result, an attacker positioned between a client and server may be able to conduct an adversary-in-the-middle "
                "(AiTM) attack by impersonating the service without possessing a valid TLS certificate. Although communications may "
                "remain encrypted, the absence of authentication undermines one of the primary security objectives of TLS: assurance of "
                "server identity.\n\n"

                "Modern clients typically disable anonymous cipher suites by default; however, their continued availability unnecessarily "
                "expands the attack surface and weakens the overall security posture of the service."
                ,
            
            "solution":
                "Disable all anonymous cipher suites and ensure the service only supports authenticated cipher suites that utilise valid "
                "TLS certificates. Specifically:\n"
                    "- Remove support for all `aNULL` cipher suites.\n"
                    "- Configure services to use modern TLS cipher suites supporting certificate-based authentication.\n"
                    "- Restrict protocol support to TLS 1.2 and TLS 1.3 where possible.\n"
                    "- Periodically review TLS configurations to ensure deprecated cipher suites are removed."
            ,

            "references": [

            ]
        },

        "Invalid TLS Certificate Configuration": {

            "title": "Invalid TLS Certificate Configuration",

            "commentary": None,

            "solution":
                "Replace the affected certificate with a valid TLS certificate that:\n"
                "- Is issued by a trusted Certificate Authority (CA) or approved internal PKI.\n"
                "- Is within its validity period and renewed before expiration.\n"
                "- Includes correct hostname information within the Subject Alternative Name (SAN) extension.\n"
                "- Presents a complete and valid certificate chain, including all required intermediate certificates.\n"
                "- Uses modern cryptographic algorithms and key lengths.\n\n"
                "Additionally, implement certificate lifecycle management processes "
                "to ensure certificates are regularly reviewed, monitored, and renewed before expiry.",

            "references": [

            ]
        }
}


#---------------------------------------
# XLSX Parsing
#---------------------------------------

def load_findings(path):

    workbook = load_workbook(path)
    sheet = workbook.active

    findings = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            plugin_id = str(int(row[1]))

        except Exception:
            continue

        findings.append(
            {
                "plugin_id": plugin_id,
                "name": row[0],
                "ip": row[11],
                "last_ip": row[12],
                "hostname": row[13],
                "system_name": row[14],
                "service": row[15],
                "severity": row[5]
            }
        )

    return findings


#---------------------------------------
# Statistics Generation
#---------------------------------------

def summarise_category(findings):

    services = set()
    host_count = count_hosts(findings)

    for finding in findings:
        services.add((finding["ip"] or finding["last_ip"], finding["service"]))

    return {"instances": len(findings), "hosts": host_count, "services": len(services)}


def count_hosts(findings):

    hosts = set()

    for finding in findings:

        if finding["ip"]:
            hosts.add(str(finding["ip"]).strip())

        if finding["last_ip"]:
            hosts.add(str(finding["last_ip"]).strip())

    return len(hosts)


#---------------------------------------
# Vulnerability Grouping
#---------------------------------------

def group_findings(findings):

    grouped = defaultdict(list)

    for finding in findings:
        plugin_id = finding["plugin_id"]

        for category, data in TLS_FINDINGS.items():
            if plugin_id in data["ids"]:
                grouped[category].append(finding)

                break

    return grouped


#---------------------------------------
# Markdown Export
#---------------------------------------

def build_certificate_commentary(findings):

    detected_plugins = {f["plugin_id"] for f in findings}
    output = []

    for plugin_id, issue in CERT_ISSUES.items():
        if plugin_id in detected_plugins:

            output.append(f"#### {issue['title']}\n")
            output.append(issue["commentary"])
            output.append("")
    
    return "\n".join(output)


def generate_markdown(grouped):

    output = []
    categories = sorted(grouped.keys())

    output.append("--------------------------------------------------")
    output.append("## TECHNICAL COMMENTARY")
    output.append("--------------------------------------------------\n")

    output.append("The assessed services were found to support a range of deprecated protocols, weak cryptographic configurations, and improperly "
        "configured TLS certificates, resulting in reduced security and trust in encrypted communications.\n")

    output.append("### Findings Overview\n")

    output.append("| Finding |Instances | Hosts | Services |")
    output.append("|---------|--------|----------|-----------|")

    for category in categories:
        stats = summarise_category(grouped[category])
        output.append(f"| {TLS_LIBRARY[category]['title']} | {stats['instances']} | {stats['hosts']} | {stats['services']} |")

    output.append("\n")

    for category in categories:
        findings = grouped[category]
        stats = summarise_category(findings)

        output.append(f"### {TLS_LIBRARY[category]['title']}\n")

        if category == "Invalid TLS Certificate Configuration":
            output.append(build_certificate_commentary(findings))

        else:
            output.append(TLS_LIBRARY[category]["commentary"])

        output.append("\n")

    output.append("--------------------------------------------------")
    output.append("## REMEDIATION")
    output.append("--------------------------------------------------\n")

    output.append("The following remediation actions are recommended to address the identified TLS-related weaknesses.\n")

    for category in categories:

        output.append(f"### {TLS_LIBRARY[category]['title']}\n")
        output.append(TLS_LIBRARY[category]["solution"])
        output.append("\n")

  
    output.append("\n### References\n")

    all_refs = set()

    for category in categories:

        for ref in TLS_LIBRARY[category]["references"]:
            all_refs.add(ref)

    for ref in sorted(all_refs):
        url = REFERENCES[ref]
        output.append(f"- [{url}]({url})")

    report_path = "tls-review.md"

    with open(report_path, "w", encoding="utf-8") as fp:
        fp.write("\n".join(output) + "\n")

    return report_path


#---------------------------------------
# Main
#---------------------------------------

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="Platform XLS export")

    args = parser.parse_args()

    findings = load_findings(args.xlsx)
    grouped = group_findings(findings)

    report = generate_markdown(grouped)

    print()
    print("[*] TLS Organiser v1.0")
    print()

    for category in sorted(grouped.keys()):

        stats = summarise_category(grouped[category])

        print(
            f"[+] {category:<40} "
            f"Hosts: {stats['hosts']:<5} "
            f"Services: {stats['services']:<5}"
        )

    print()
    print(f"[+] Findings Processed : {len(findings)}")
    print(f"[+] Root Causes        : {len(grouped)}")
    print(f"[+] Output File        : {report}")
    print()


if __name__ == "__main__":
    main()